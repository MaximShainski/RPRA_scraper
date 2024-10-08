from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium_recaptcha_solver import RecaptchaSolver
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook
from datetime import datetime
import numpy
import os
import re

def checkChanges():

    test_ua = 'Mozilla/5.0 (Windows NT 4.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2049.0 Safari/537.36'

    options = Options()
    options.add_argument("start-maximized")
    options.add_argument("disable-infobars")

    options.add_argument("--window-size=1920,1080")

    options.add_argument(f'--user-agent={test_ua}')

    options.add_argument('--no-sandbox')
    options.add_argument("--disable-extensions")

    workbook = load_workbook('ES Registry Database - Updated.xlsx')
    worksheet = workbook['Reuse']

    print(worksheet.max_row)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    #solver = RecaptchaSolver(driver=driver)

    driver.get('https://excesssoilnotices.rpra.ca/s/?language=en_US')   
    wait = WebDriverWait(driver, 5)
    name = input("Click enter when done the captcha")
    try:
        arrowEnabled = True
        arrowCounter = -1
        while (arrowEnabled):
            counter = -1
            arrowCounter += 1
            time.sleep(numpy.random.uniform(1,2))
            #arrows = driver.find_elements(By.XPATH, '//button[@class="slds-button slds-button_outline-brand buttonSize textFont textColour" and @type="button"]')
            nextArrow = driver.find_element(By.XPATH, '(//button[@class="slds-button slds-button_outline-brand buttonSize textFont textColour" and @type="button"])[3]')
            arrowEnabled = nextArrow.is_enabled()
            municipalities = driver.find_elements(By.XPATH, '//span[contains(@class, "uiOutputText") and substring(text(), string-length(text()) - 2) = " of"]')

            while counter < len(municipalities) - 1:
                for n in range(arrowCounter):
                    time.sleep(numpy.random.uniform(1,2))
                    nextArrow = driver.find_element(By.XPATH, '(//button[@class="slds-button slds-button_outline-brand buttonSize textFont textColour" and @type="button"])[3]')
                    try:
                        nextArrow.click()
                    except:
                        input("Solve the captcha and then click enter in the terminal")
                time.sleep(numpy.random.uniform(1,2))
                counter += 1
                elements = driver.find_elements(By.XPATH, '//*[@data-aura-class="cRegistryPublicPortalFilingAction"]')
                municipalities = driver.find_elements(By.XPATH, '//span[contains(@class, "uiOutputText") and substring(text(), string-length(text()) - 2) = " of"]')
                notice = driver.find_elements(By.XPATH, '//span[contains(@class, "uiOutputText") and (text()="RS" or text()="PA")]')
                if notice[counter].text == "RS":
                    worksheet = workbook['Reuse']
                elif notice[counter].text == "PA":
                    worksheet = workbook['Project Area']
                else:
                    continue
                municipality = municipalities[counter]
        
                if worksheet == workbook['Reuse']:
                    row = worksheet.max_row + 1

                    worksheet['F' + str(row)] = municipality.text.split(",")[0]
                    try:
                        elements[counter].click()
                    except:
                        input("Solve the captcha and then click enter in the terminal (2)")
                        elements[counter].click()

                    time.sleep(numpy.random.uniform(2,4))

                    projectName = driver.find_element(By.XPATH, '//*[starts-with(@data-aura-rendered-by, "225")]')
                    worksheet['A' + str(row)] = projectName.text

                    companyName = driver.find_element(By.XPATH, '//*[starts-with(@data-aura-rendered-by, "40:") and @class = "slds-cell-wrap"]')
                    worksheet['B' + str(row)] = companyName.text

                    location = driver.find_element(By.XPATH, '//span[contains(@data-aura-rendered-by, "24:") and @class="uiOutputText" and contains(., ",")]')
                    worksheet['C' + str(row)] = location.text

                    table = driver.find_element(By.XPATH, '//span[contains(@data-aura-rendered-by, "47") and contains (., "-")]')
                    worksheet['D' + str(row)] = table.text

                    soil = driver.find_element(By.XPATH, '//*[contains(@data-aura-rendered-by, "729")]')
                    worksheet['E' + str(row)] = soil.text

                    dateAdded = driver.find_element(By.XPATH, '//lightning-formatted-date-time[starts-with(@data-aura-rendered-by, "53")]')
                    cleanDate = dateAdded.text
                    formattedDate = datetime.strptime(cleanDate, '%d-%b-%Y') # Convert datetime object to date
                    worksheet['G' + str(row)] = formattedDate.strftime('%Y-%m-%d') # Format as string without time

                    contactName = driver.find_element(By.XPATH, '//span[contains(@data-aura-rendered-by, "24:") and @class="uiOutputText" and not(contains(., ","))]')
                    contactMail = driver.find_element(By.XPATH, '//span[contains(@data-aura-rendered-by, "87")]')
                    worksheet['H' + str(row)] = contactName.text + " " + contactMail.text

                    #url = driver.current_url
                    worksheet['I' + str(row)] = driver.current_url

                    coordinates = driver.find_element(By.XPATH, '//span[contains(@data-aura-rendered-by, "46") and @class="uiOutputText" and contains(., ",") and contains(translate(., "0123456789,-", ""), "")]')
                    worksheet['J' + str(row)] = coordinates.text

                elif (worksheet == workbook['Project Area']):
                    row = worksheet.max_row

                    #worksheet['F' + str(row)] = municipality.text.split(",")[0]
                    try:
                        elements[counter].click()
                    except:
                        input("Solve the captcha and then click enter in the terminal (3)")

                    time.sleep(numpy.random.uniform(2,4))
                    
                    siteNames = driver.find_elements(By.XPATH, "//div[@class='slds-cell-wrap' and text()='Site Name']/following::td[1]")
                    for item in siteNames:
                        row += 1
                        projectName = driver.find_element(By.XPATH, '//*[contains(@data-aura-rendered-by, "228")]')
                        worksheet['A' + str(row)] = projectName.text

                        companyName = driver.find_element(By.XPATH, '//*[starts-with(@data-aura-rendered-by, "40:") and @class = "slds-cell-wrap"]')
                        worksheet['B' + str(row)] = companyName.text
                        try:
                            driver.find_element(By.XPATH, '//*[contains(@data-aura-rendered-by, "316") and @class = "slds-truncate"]')
                            location = driver.find_element(By.XPATH, '//lightning-formatted-rich-text[contains(@data-aura-rendered-by, "323")]/span[@part="formatted-rich-text"]')
                        except:
                            location = driver.find_elements(By.XPATH, '//*[self::lightning-formatted-rich-text/span[@part="formatted-rich-text" and normalize-space(.)] or self::span[@part="formatted-rich-text" and normalize-space(.)]]')
                            location = location[5]
                        worksheet['C' + str(row)] = location.text

                        city = driver.find_element(By.XPATH, "//div[text()='Municipality']/following::td[1]")
                        worksheet['D' + str(row)] = city.text.split(",")[0]

                        site = item.text
                        worksheet['E' + str(row)] = site

                        soil = driver.find_element(By.XPATH, '//*[text()="' + site + '"]/following::td[contains(., "Estimated Amount of Soil (m3)")]/following-sibling::td[1]')
                        worksheet['F' + str(row)] = soil.text

                        dateAdded = driver.find_element(By.XPATH, '//lightning-formatted-date-time[starts-with(@data-aura-rendered-by, "53")]')
                        cleanDate = dateAdded.text #Might have to change depending on whetehr it's date submitted or modified, right now it's submitted
                        formattedDate = datetime.strptime(cleanDate, '%d-%b-%Y') # Convert datetime object to date
                        worksheet['H' + str(row)] = formattedDate.strftime('%Y-%m-%d') # Format as string without time     

                        operator = driver.find_element(By.XPATH, "//span[@class='uiOutputEmail' and contains(@data-aura-rendered-by, '197')]/a")
                        worksheet['I' + str(row)] = operator.text

                        qualifiedCompany = driver.find_elements(By.XPATH, "//div[text()='Company Name']/following::td[1]")
                        worksheet['J' + str(row)] = qualifiedCompany[1].text

                        qualifiedPersonName = driver.find_element(By.XPATH, "//div[text()='Contact Name']/following::td[1]")
                        qualifiedPersonContact = driver.find_element(By.XPATH, "//div[text()='Email']/following::td[1]")
                        worksheet['K' + str(row)] = qualifiedPersonName.text + " " + qualifiedPersonContact.text

                        worksheet['L' + str(row)] = driver.current_url

                        latitude = driver.find_element(By.XPATH, '//*[text()="' + site + '"]/following::td[contains(., "Latitude")]/following-sibling::td[1]')
                        worksheet['M' + str(row)] = latitude.text

                        latitude = driver.find_element(By.XPATH, '//*[text()="' + site + '"]/following::td[contains(., "Latitude")]/following-sibling::td[1]')
                        worksheet['N' + str(row)] = latitude.text

                driver.execute_script("window.history.go(-1)")
    finally:
        workbook.save(filename="test.xlsx")
        driver.quit()
        

checkChanges()